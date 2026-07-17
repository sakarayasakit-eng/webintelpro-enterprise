import json
import os

from engine import AnalysisEngine
from reporter import ReportGenerator


def _result(html, headers, cookies):
    return AnalysisEngine().analyze("https://acme.example", html, headers, cookies)


def test_excel_export(tmp_path, wordpress_html, wordpress_headers, wordpress_cookies):
    from openpyxl import load_workbook
    result = _result(wordpress_html, wordpress_headers, wordpress_cookies)
    path = tmp_path / "r.xlsx"
    ReportGenerator().save_excel(result, str(path))
    assert path.exists() and path.stat().st_size > 0
    wb = load_workbook(str(path))
    assert set(wb.sheetnames) == {"Summary", "Technologies", "Recommendations"}


def test_pdf_export(tmp_path, wordpress_html, wordpress_headers, wordpress_cookies):
    result = _result(wordpress_html, wordpress_headers, wordpress_cookies)
    path = tmp_path / "r.pdf"
    ReportGenerator().save_pdf(result, str(path))
    assert path.exists()
    with open(path, "rb") as f:
        assert f.read(5) == b"%PDF-"


def test_cache_roundtrip(tmp_path, wordpress_html, wordpress_headers):
    from core.cache import CrawlCache
    from crawler import CrawlResult
    cache = CrawlCache(directory=str(tmp_path), ttl=3600)
    cr = CrawlResult(url="https://c.example", final_url="https://c.example",
                     status_code=200, html=wordpress_html, headers=wordpress_headers,
                     cookies={}, elapsed=0.1)
    assert cache.get("https://c.example") is None
    cache.put(cr)
    got = cache.get("https://c.example")
    assert got is not None and got.status_code == 200 and got.html == wordpress_html
