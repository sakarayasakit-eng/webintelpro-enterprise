"""
Offline tests for Phase 2E Authentication & Identity Intelligence.

This sub-system is passive-only (HTML/JS/headers/cookies already fetched;
no browser automation, no login, no credentials, no brute force, no extra
requests), so every test here is offline by construction. Coverage follows
the Phase 2E contract: provider/identity-provider/protocol/security-feature
signature matching (including the newly-added cookie-name source), the
guarantees that matter more than any single detection (the default
detect() path is unchanged, header *and* cookie values are never inspected,
generic prose does not false-trigger), confidence scoring, and graceful
degradation.
"""

from __future__ import annotations

import json

import pytest

from modules.auth_detection import AuthCategory, AuthDetector
from modules.auth_detection.detector import ALL_SIGNATURES
from technology.detector import TechnologyDetector
from technology.parser import HTMLParser


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def analyze(html: str, headers=None, cookies=None, url: str = "https://site.test/"):
    parsed = HTMLParser().parse(html)
    return AuthDetector().analyze(parsed, url, headers, cookies)


def names(analysis):
    return {t.name for t in analysis.technologies}


def page(body: str, head: str = "") -> str:
    return f"<html><head>{head}</head><body>{body}</body></html>"


def script(js: str) -> str:
    return page(f"<script>{js}</script>")


# ---------------------------------------------------------------------------
# positive detections across categories
# ---------------------------------------------------------------------------

def test_detects_a_provider_from_sdk_and_domain():
    html = script('new auth0.WebAuth({domain: "my-tenant.auth0.com"}); '
                  'loginWithRedirect();')
    analysis = analyze(html)
    assert "Auth0" in names(analysis)
    finding = next(f for f in analysis.report.findings if f.name == "Auth0")
    assert finding.category is AuthCategory.PROVIDER


def test_detects_an_identity_provider():
    html = script('const auth = new OktaAuth({issuer: "https://dev-123.okta.com"});')
    analysis = analyze(html)
    assert "Okta" in names(analysis)
    finding = next(f for f in analysis.report.findings if f.name == "Okta")
    assert finding.category is AuthCategory.IDENTITY_PROVIDER


def test_detects_azure_entra_id():
    html = script('const msal = new PublicClientApplication({auth: '
                  '{authority: "https://login.microsoftonline.com/tenant"}});')
    analysis = analyze(html)
    assert "Azure Entra ID" in names(analysis)


def test_detects_keycloak():
    html = script('const kc = new Keycloak("/auth/realms/myrealm/protocol/'
                  'openid-connect/auth");')
    analysis = analyze(html)
    assert "Keycloak" in names(analysis)


@pytest.mark.parametrize("snippet,expected", [
    ('fetch("/oauth2/token", {body: "grant_type=authorization_code"});', "OAuth2"),
    ('fetch("/.well-known/openid-configuration");', "OpenID Connect"),
    ('const req = "SAMLRequest=" + encodeURIComponent(xml);', "SAML"),
    ('const code_verifier = generateVerifier(); const code_challenge = hash(v);', "PKCE"),
])
def test_detects_protocol(snippet, expected):
    analysis = analyze(script(snippet))
    assert expected in names(analysis)
    finding = next(f for f in analysis.report.findings if f.name == expected)
    assert finding.category is AuthCategory.PROTOCOL


def test_detects_jwt_via_quote_anchored_prefix():
    """Regression guard: the bare 'eyJhbGciOiJ' prefix is pure-alphanumeric,
    so the rule engine's word-boundary matcher would never match it inside a
    real JWT (which always continues with more base64url characters
    immediately after). The signature must anchor to a preceding quote/space
    instead, which forces substring matching."""
    html = script('const jwt = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.payload.sig";')
    analysis = analyze(html)
    assert "JWT" in names(analysis)


def test_jwt_bare_prefix_pattern_is_not_used():
    """The pure-alphanumeric 'eyJhbGciOiJ' (no anchor) must never be a
    pattern on its own -- it would silently never match real JWTs. Guards
    against a future edit reintroducing the bug fixed above."""
    jwt_sig = next(s for s in ALL_SIGNATURES if s.name == "JWT")
    assert "eyJhbGciOiJ" not in jwt_sig.patterns


def test_detects_security_feature_csrf():
    html = script('const token = document.querySelector(\'meta[name="csrf-token"]\');')
    analysis = analyze(html)
    assert "CSRF Protection" in names(analysis)


def test_detects_security_feature_refresh_tokens():
    html = script('fetch("/token", {body: "grant_type=refresh_token"});')
    analysis = analyze(html)
    assert "Refresh Tokens" in names(analysis)


def test_detects_token_storage():
    html = script("localStorage.setItem('access_token', token);")
    analysis = analyze(html)
    assert "Token Storage" in names(analysis)


def test_detects_login_popup():
    html = script("await client.loginWithPopup();")
    analysis = analyze(html)
    assert "Login Popup" in names(analysis)


def test_detects_silent_auth():
    html = script("const token = await client.getTokenSilently();")
    analysis = analyze(html)
    assert "Silent Auth" in names(analysis)


def test_detects_mfa():
    html = script('console.log("otpauth://totp/Example:user?secret=ABC");')
    analysis = analyze(html)
    assert "MFA" in names(analysis)


def test_detects_passkeys_plural_form():
    """Regression guard: 'passkey' alone is pure-alphanumeric, so it would
    never match the far more common plural 'passkeys' (trailing 's' blocks
    the word boundary). Both forms must be registered."""
    html = page("<p>This site supports passkeys for passwordless login.</p>")
    analysis = analyze(html)
    assert "Passkeys / WebAuthn" in names(analysis)


def test_detects_webauthn_api_calls():
    html = script("navigator.credentials.get({publicKey: options});")
    analysis = analyze(html)
    assert "Passkeys / WebAuthn" in names(analysis)


# ---------------------------------------------------------------------------
# header-name and cookie-name evidence
# ---------------------------------------------------------------------------

def test_detects_provider_via_response_header_name_only():
    analysis = analyze(page("<p>hi</p>"), headers={"x-clerk-auth-status": "signed-in"})
    assert "Clerk" in names(analysis)
    finding = next(f for f in analysis.report.findings if f.name == "Clerk")
    assert finding.evidence == ["x-clerk-auth-status"]


def test_detects_session_cookie_via_cookie_name_only():
    analysis = analyze(page("<p>hi</p>"), cookies={"connect.sid": "s%3AsuperSecretValue"})
    assert "Session Cookies" in names(analysis)
    finding = next(f for f in analysis.report.findings if f.name == "Session Cookies")
    assert finding.evidence == ["connect.sid"]
    assert finding.category is AuthCategory.SECURITY_FEATURE


def test_detects_clerk_via_distinctive_cookie_name():
    analysis = analyze(page("<p>hi</p>"), cookies={"__client_uat": "1234567890"})
    assert "Clerk" in names(analysis)


def test_cookie_matching_is_case_sensitive():
    """Cookie names are case-sensitive per RFC 6265 (unlike HTTP header field
    names, which are case-insensitive and are lowercased before matching).
    A cookie named with different casing than the real framework default
    must not match."""
    analysis = analyze(page("<p>hi</p>"), cookies={"CONNECT.SID": "x"})
    assert "Session Cookies" not in names(analysis)


# ---------------------------------------------------------------------------
# mixed stack
# ---------------------------------------------------------------------------

def test_mixed_auth_stack_reports_every_category_present():
    html = script(
        'new auth0.WebAuth({domain: "t.auth0.com"}); '
        'const c = new OktaAuth({issuer: "https://dev.okta.com"}); '
        'fetch("/oauth2/token", {body: "grant_type=authorization_code"}); '
        'localStorage.setItem("access_token", t);'
    )
    analysis = analyze(html)
    found = names(analysis)
    assert {"Auth0", "Okta", "OAuth2", "Token Storage"} <= found
    grouped = analysis.report.to_dict()
    assert len(grouped["providers"]) == 1
    assert len(grouped["identity_providers"]) == 1
    assert len(grouped["protocols"]) == 1
    assert len(grouped["security_features"]) == 1


# ---------------------------------------------------------------------------
# negative detections / empty pages
# ---------------------------------------------------------------------------

def test_empty_page_yields_no_findings():
    analysis = analyze(page(""))
    assert analysis.technologies == []
    assert analysis.report.findings == []
    assert analysis.report.errors == []


def test_unrelated_page_yields_no_findings():
    html = page("<h1>Welcome to our bakery</h1><p>Fresh bread daily.</p>")
    analysis = analyze(html)
    assert analysis.technologies == []


# ---------------------------------------------------------------------------
# false-positive prevention
# ---------------------------------------------------------------------------

def test_generic_prose_does_not_false_trigger():
    """Ordinary sentences using the words 'session', 'token', 'login',
    'secure', 'auth' on their own must not fire -- every real signature
    pattern is qualified (SDK/host/protocol-specific)."""
    html = page("<p>Please login to start your session. Keep your account "
               "secure and never share your token or auth code.</p>")
    analysis = analyze(html)
    assert names(analysis) == set()


def test_every_signature_pattern_is_qualified_not_a_bare_word():
    """Precision-over-recall guarantee at the data level: no pattern is a
    single plain dictionary/generic-auth word with no vendor/protocol/API
    qualifier that could match ordinary prose."""
    bare_generic_words = {
        "session", "token", "login", "auth", "secure", "password",
        "cookie", "oauth", "saml", "jwt", "sso", "mfa", "otp",
        "okta", "clerk", "keycloak", "stytch", "descope",
    }
    for sig in ALL_SIGNATURES:
        for pattern in sig.patterns:
            assert pattern.lower().strip() not in bare_generic_words, (
                f"{sig.name!r} has an unqualified bare-word pattern {pattern!r}")


def test_header_values_are_never_inspected_only_names():
    """A secret-looking header *value* must never itself cause a match --
    only the header *name* is compared against signature.headers."""
    analysis = analyze(page("<p>hi</p>"),
                       headers={"X-Custom": "clerk secret sk_live_abcdef"})
    assert names(analysis) == set()


def test_cookie_values_are_never_inspected_only_names():
    """A JWT-shaped or secret-looking cookie *value* must never itself cause
    a match -- only the cookie *name* is compared against signature.cookies.
    This is Phase 2E's new invariant, mirroring the header guarantee above."""
    analysis = analyze(page("<p>hi</p>"),
                       cookies={"my_custom_cookie": "eyJhbGciOiJIUzI1NiJ9.fake.sig"})
    assert names(analysis) == set()


def test_no_duplicate_signature_names():
    seen = {}
    for sig in ALL_SIGNATURES:
        assert sig.name not in seen, (
            f"{sig.name!r} is defined twice: {sig.category} and {seen[sig.name]}")
        seen[sig.name] = sig.category


# ---------------------------------------------------------------------------
# confidence scoring
# ---------------------------------------------------------------------------

def test_text_only_evidence_uses_auth_evidence_weight():
    analysis = analyze(script('fetch("/.well-known/openid-configuration");'))
    finding = next(f for f in analysis.report.findings if f.name == "OpenID Connect")
    assert 0.30 <= finding.confidence < 0.60


def test_combining_header_and_text_evidence_raises_confidence():
    html = script('window.Clerk = {};')
    analysis = analyze(html, headers={"x-clerk-auth-status": "signed-in"})
    finding = next(f for f in analysis.report.findings if f.name == "Clerk")
    text_only = analyze(html)
    text_only_finding = next(f for f in text_only.report.findings if f.name == "Clerk")
    assert finding.confidence > text_only_finding.confidence


def test_low_confidence_findings_are_not_reported():
    from modules.auth_detection import AuthDetectionConfig
    parsed = HTMLParser().parse(script('fetch("/oauth2/token");'))
    analysis = AuthDetector(config=AuthDetectionConfig(min_confidence=0.99)).analyze(
        parsed, "https://site.test/")
    assert analysis.technologies == []


# ---------------------------------------------------------------------------
# report shape
# ---------------------------------------------------------------------------

def test_report_to_dict_has_every_category_key():
    grouped = analyze(page("<p>hi</p>")).report.to_dict()
    for cat in AuthCategory:
        assert cat.value in grouped
        assert grouped[cat.value] == []
    assert grouped["total_findings"] == 0


# ---------------------------------------------------------------------------
# TechnologyDetector integration
# ---------------------------------------------------------------------------

def test_default_detect_path_is_unaffected():
    html = script('new auth0.WebAuth({domain: "t.auth0.com"});')
    detector = TechnologyDetector()
    report = detector.detect("https://site.test/", html=html, cookies={"connect.sid": "x"})
    assert report.authentication is None
    assert "authentication" not in report.to_dict()
    assert "Auth0" not in {t.name for t in report.technologies}
    assert "Session Cookies" not in {t.name for t in report.technologies}


def test_opt_in_detect_path_populates_report():
    html = script('new auth0.WebAuth({domain: "t.auth0.com"}); loginWithRedirect();')
    detector = TechnologyDetector()
    report = detector.detect("https://site.test/", html=html,
                             cookies={"connect.sid": "x"}, analyze_auth=True)
    assert report.authentication is not None
    assert report.authentication["total_findings"] >= 1
    assert any(t.category == "auth_providers" for t in report.technologies)
    assert any(t.category == "auth_security_features" for t in report.technologies)
    assert "authentication" in report.to_dict()


def test_auth_detection_failure_degrades_gracefully(monkeypatch):
    detector = TechnologyDetector()

    class _Boom:
        def analyze(self, *a, **k):
            raise RuntimeError("boom")

    monkeypatch.setattr(detector, "_auth_detector", _Boom())
    report = detector.detect("https://site.test/", html=page("<p>hi</p>"),
                             analyze_auth=True)
    assert report.authentication is None
    assert report.total_detected == len(report.technologies)  # did not crash


# ---------------------------------------------------------------------------
# reporter / exporter integration
# ---------------------------------------------------------------------------

def _full_result(analyze_auth: bool):
    html = script('new auth0.WebAuth({domain: "t.auth0.com"}); loginWithRedirect();')
    detector = TechnologyDetector()
    report = detector.detect("https://site.test/", html=html,
                             cookies={"connect.sid": "x"}, debug=True,
                             analyze_auth=analyze_auth)
    return {
        "url": "https://site.test/", "technology": report,
        "overall": {"score": 80, "grade": "B",
                    "parts": {"seo": 80, "security": 80, "performance": 80,
                              "accessibility": 80}},
        "seo": {"score": 80, "grade": "B", "issues": []},
        "security": {"score": 80, "grade": "B", "issues": []},
        "performance": {"score": 80, "grade": "B", "issues": []},
        "accessibility": {"score": 80, "grade": "B", "issues": []},
        "recommendations": [], "recommendation_summary": {"by_severity": {}},
    }


def test_console_authentication_section_only_appears_when_enabled():
    from reporter import ReportGenerator
    rg = ReportGenerator()
    off = rg.console_str(_full_result(analyze_auth=False))
    on = rg.console_str(_full_result(analyze_auth=True))
    assert "AUTHENTICATION" not in off
    assert "AUTHENTICATION" in on
    assert "Auth0" in on


def test_json_export_includes_authentication_only_when_enabled():
    from reporter import ReportGenerator
    rg = ReportGenerator()
    off = json.loads(rg.to_json(_full_result(analyze_auth=False)))
    on = json.loads(rg.to_json(_full_result(analyze_auth=True)))
    assert "authentication" not in off["technology"]
    assert "authentication" in on["technology"]
    assert on["technology"]["authentication"]["total_findings"] >= 1


def test_html_export_includes_authentication_section_only_when_enabled():
    from reporter import ReportGenerator
    rg = ReportGenerator()
    off = rg.to_html(_full_result(analyze_auth=False))
    on = rg.to_html(_full_result(analyze_auth=True))
    assert "Authentication (" not in off
    assert "Authentication (" in on


def test_excel_export_adds_authentication_sheet_only_when_enabled(tmp_path):
    from openpyxl import load_workbook
    from reporter import ReportGenerator
    rg = ReportGenerator()

    off_path = tmp_path / "off.xlsx"
    rg.save_excel(_full_result(analyze_auth=False), str(off_path))
    assert "Authentication" not in load_workbook(str(off_path)).sheetnames

    on_path = tmp_path / "on.xlsx"
    rg.save_excel(_full_result(analyze_auth=True), str(on_path))
    assert "Authentication" in load_workbook(str(on_path)).sheetnames


def test_pdf_export_succeeds_with_authentication_enabled(tmp_path):
    from reporter import ReportGenerator
    rg = ReportGenerator()
    path = tmp_path / "auth.pdf"
    rg.save_pdf(_full_result(analyze_auth=True), str(path))
    assert path.exists()
    with open(path, "rb") as f:
        assert f.read(5) == b"%PDF-"
