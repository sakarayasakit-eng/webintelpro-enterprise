"""
WebIntelPro Enterprise X
Report Generator

Renders analysis results to console, JSON, standalone HTML, Excel, and PDF.
"""

from __future__ import annotations

import html as _html
import json
from datetime import datetime

_SEV_ORDER = ["critical", "high", "medium", "low"]
_SEV_COLOR = {"critical": "#dc2626", "high": "#ea580c",
              "medium": "#ca8a04", "low": "#65a30d"}

# Phase 2C: display order/labels for modules.api_discovery.ApiKind buckets,
# mirroring ApiDiscoveryReport.to_dict()'s grouped keys.
_API_KIND_ORDER = ["rest", "graphql", "openapi", "websocket", "sse",
                   "json", "jsonrpc", "xmlrpc", "trpc", "grpc_web"]
_API_KIND_LABELS = {
    "rest": "REST", "graphql": "GraphQL", "openapi": "Swagger/OpenAPI",
    "websocket": "WebSocket", "sse": "SSE", "json": "JSON",
    "jsonrpc": "JSON-RPC", "xmlrpc": "XML-RPC", "trpc": "tRPC",
    "grpc_web": "gRPC-Web",
}

# Phase 2D: display order/labels for modules.ai_detection.AICategory buckets,
# mirroring AIDetectionReport.to_dict()'s grouped keys.
_AI_CATEGORY_ORDER = ["providers", "open_source_models", "local_ai",
                      "frameworks", "sdks", "vector_databases",
                      "embedding_services", "infrastructure"]
_AI_CATEGORY_LABELS = {
    "providers": "Providers", "open_source_models": "Open Source Models",
    "local_ai": "Local AI", "frameworks": "Frameworks", "sdks": "SDKs",
    "vector_databases": "Vector Databases",
    "embedding_services": "Embedding Services",
    "infrastructure": "Infrastructure",
}

# Phase 2E: display order/labels for modules.auth_detection.AuthCategory
# buckets, mirroring AuthDetectionReport.to_dict()'s grouped keys.
_AUTH_CATEGORY_ORDER = ["providers", "identity_providers", "protocols",
                        "security_features"]
_AUTH_CATEGORY_LABELS = {
    "providers": "Providers", "identity_providers": "Identity Providers",
    "protocols": "Protocols", "security_features": "Security Features",
}


class ReportGenerator:

    # ---------------------------------------------------------------- console
    def console(self, result: dict) -> None:
        print(self.console_str(result))

    def console_str(self, result: dict) -> str:
        tech = result["technology"]
        overall = result["overall"]
        crawl = result.get("crawl")
        L = []
        bar = "=" * 64
        L.append(bar)
        L.append("  WebIntelPro Enterprise X  -  Website Intelligence Report")
        L.append(bar)
        L.append(f"URL          : {result['url']}")
        if crawl is not None:
            L.append(f"Status       : {crawl.status_code}   ({crawl.elapsed:.2f}s)")
        L.append(f"Server       : {tech.server or 'Unknown'}")
        L.append(f"CMS          : {tech.cms or 'None detected'}")
        L.append("")
        L.append(f"OVERALL SCORE: {overall['score']}/100  (grade {overall['grade']})")
        p = overall["parts"]
        L.append(f"   SEO {p['seo']:>3}  |  Security {p['security']:>3}  "
                 f"|  Performance {p['performance']:>3}  |  A11y {p['accessibility']:>3}")

        L.append("")
        L.append("-" * 64)
        L.append(f"DETECTED TECHNOLOGIES ({tech.total_detected})")
        L.append("-" * 64)
        if tech.technologies:
            for cat, names in sorted(tech.by_category().items()):
                L.append(f"  {cat:12}: {', '.join(names)}")
        else:
            L.append("  None detected.")

        debug_mode = any(t.debug for t in tech.technologies)
        if debug_mode:
            L.append("")
            L.append("-" * 64)
            L.append("DETECTION EVIDENCE (debug)")
            L.append("-" * 64)
            for t in tech.technologies:
                if not t.debug:
                    continue
                ver = f"  version {t.version}" if t.version else ""
                L.append(f"  {t.name}  [{t.category}]  confidence {t.confidence:.2f}{ver}")
                for source, info in sorted(t.debug.items()):
                    extra = f"  (+{info['extra_matches']} more)" if info["extra_matches"] else ""
                    patterns = ", ".join(info["patterns"][:5])
                    L.append(f"      {source:12} weight {info['weight']:.2f}  "
                             f"matched: {patterns}{extra}")

        api = tech.api_discovery
        if api is not None:
            L.append("")
            L.append("-" * 64)
            flags = []
            if api.get("probed"):
                flags.append("probed")
            if api.get("truncated"):
                flags.append("truncated")
            suffix = f"  ({', '.join(flags)})" if flags else ""
            L.append(f"API DISCOVERY ({api['total_findings']}){suffix}")
            L.append("-" * 64)
            if api["total_findings"]:
                for kind in _API_KIND_ORDER:
                    findings = api.get(kind) or []
                    if not findings:
                        continue
                    L.append(f"  {_API_KIND_LABELS.get(kind, kind)} ({len(findings)}):")
                    for f in findings:
                        reach = ""
                        if f.get("reachable") is not None:
                            status = f" {f['http_status']}" if f.get("http_status") else ""
                            reach = ("  [reachable" if f["reachable"] else "  [unreachable") \
                                    + f"{status}]"
                        L.append(f"    - {f['name']}  confidence {f['confidence']:.2f}"
                                 f"  source {f['source']}{reach}")
                        if debug_mode:
                            for ev in f.get("evidence", []):
                                L.append(f"        evidence: {ev}")
            else:
                L.append("  No API surface discovered.")
            if api.get("errors"):
                L.append(f"  errors: {', '.join(api['errors'])}")

        ai_stack = tech.ai_stack
        if ai_stack is not None:
            L.append("")
            L.append("-" * 64)
            flags = ["truncated"] if ai_stack.get("truncated") else []
            suffix = f"  ({', '.join(flags)})" if flags else ""
            L.append(f"AI STACK ({ai_stack['total_findings']}){suffix}")
            L.append("-" * 64)
            if ai_stack["total_findings"]:
                for cat in _AI_CATEGORY_ORDER:
                    findings = ai_stack.get(cat) or []
                    if not findings:
                        continue
                    L.append(f"  {_AI_CATEGORY_LABELS.get(cat, cat)} ({len(findings)}):")
                    for f in findings:
                        ver = f"  v{f['version']}" if f.get("version") else ""
                        L.append(f"    - {f['name']}  confidence {f['confidence']:.2f}"
                                 f"{ver}  source {'+'.join(f['sources'])}")
                        if debug_mode:
                            for ev in f.get("evidence", []):
                                L.append(f"        evidence: {ev}")
            else:
                L.append("  No AI stack detected.")
            if ai_stack.get("errors"):
                L.append(f"  errors: {', '.join(ai_stack['errors'])}")

        auth = tech.authentication
        if auth is not None:
            L.append("")
            L.append("-" * 64)
            flags = ["truncated"] if auth.get("truncated") else []
            suffix = f"  ({', '.join(flags)})" if flags else ""
            L.append(f"AUTHENTICATION ({auth['total_findings']}){suffix}")
            L.append("-" * 64)
            if auth["total_findings"]:
                for cat in _AUTH_CATEGORY_ORDER:
                    findings = auth.get(cat) or []
                    if not findings:
                        continue
                    L.append(f"  {_AUTH_CATEGORY_LABELS.get(cat, cat)} ({len(findings)}):")
                    for f in findings:
                        ver = f"  v{f['version']}" if f.get("version") else ""
                        L.append(f"    - {f['name']}  confidence {f['confidence']:.2f}"
                                 f"{ver}  source {'+'.join(f['sources'])}")
                        if debug_mode:
                            for ev in f.get("evidence", []):
                                L.append(f"        evidence: {ev}")
            else:
                L.append("  No authentication architecture detected.")
            if auth.get("errors"):
                L.append(f"  errors: {', '.join(auth['errors'])}")

        def section(title, data):
            L.append("")
            L.append("-" * 64)
            L.append(f"{title}   {data['score']}/100  (grade {data['grade']})")
            L.append("-" * 64)
            for issue in data["issues"] or ["No issues found."]:
                L.append(f"  [!] {issue}" if data["issues"] else "  " + issue)

        section("TECHNICAL SEO", result["seo"])
        section("SECURITY", result["security"])
        section("PERFORMANCE", result["performance"])
        section("ACCESSIBILITY", result["accessibility"])

        v = result.get("vitals")
        if v:
            L.append("")
            L.append("-" * 64)
            L.append("CORE WEB VITALS")
            L.append("-" * 64)
            if not v.get("available"):
                L.append(f"  unavailable - {v.get('reason','')}")
            else:
                for k, m in v["metrics"].items():
                    unit = "" if k == "cls" else "ms"
                    L.append(f"  {k.upper():5}: {m['value']}{unit}  ({m['rating']})")

        site = result.get("site")
        if site:
            L.append("")
            L.append("-" * 64)
            L.append("SITE CHECKS")
            L.append("-" * 64)
            rb = site.get("robots", {}); sm = site.get("sitemap", {}); tls = site.get("tls", {})
            L.append(f"  robots.txt : {'found' if rb.get('exists') else 'missing'}"
                     f" ({rb.get('disallow', 0)} disallow rules)")
            L.append(f"  sitemap    : {'found' if sm.get('exists') else 'missing'}"
                     f" ({sm.get('urls', 0)} URLs)")
            if tls.get("checked"):
                exp = tls.get("expires_in_days")
                L.append(f"  TLS        : {tls.get('protocol','')}"
                         + (f", cert expires in {exp} days" if exp is not None else ""))

        recs = result.get("recommendations", [])
        L.append("")
        L.append("=" * 64)
        summ = result.get("recommendation_summary", {}).get("by_severity", {})
        L.append("PRIORITIZED RECOMMENDATIONS  "
                 f"(critical {summ.get('critical',0)}, high {summ.get('high',0)}, "
                 f"medium {summ.get('medium',0)}, low {summ.get('low',0)})")
        L.append("=" * 64)
        if recs:
            for r in recs:
                L.append(f"  [{r['severity'].upper():8}] {r['area']}: {r['issue']}")
                L.append(f"             -> {r['recommendation']}")
        else:
            L.append("  Nothing to fix - great job!")
        L.append(bar)
        return "\n".join(L)

    # ------------------------------------------------------------------- json
    def to_json(self, result: dict) -> str:
        payload = {
            "url": result["url"],
            "generated": datetime.utcnow().isoformat() + "Z",
            "overall": result["overall"],
            "recommendation_summary": result.get("recommendation_summary"),
            "recommendations": result.get("recommendations", []),
            "technology": result["technology"].to_dict(),
            "seo": result["seo"],
            "security": result["security"],
            "performance": result["performance"],
            "accessibility": result["accessibility"],
        }
        crawl = result.get("crawl")
        if crawl is not None:
            payload["crawl"] = {"final_url": crawl.final_url,
                                "status_code": crawl.status_code,
                                "elapsed": crawl.elapsed}
        return json.dumps(payload, indent=2)

    def save_json(self, result: dict, path: str) -> None:
        with open(path, "w", encoding="utf-8") as f:
            f.write(self.to_json(result))

    # ------------------------------------------------------------------- html
    def to_html(self, result: dict) -> str:
        tech = result["technology"]
        overall = result["overall"]
        esc = lambda x: _html.escape(str(x))

        def gcolor(g):
            return {"A": "#16a34a", "B": "#65a30d", "C": "#ca8a04",
                    "D": "#ea580c", "F": "#dc2626"}.get(g, "#6b7280")

        def card(title, data):
            issues = "".join(f"<li>{esc(i)}</li>" for i in data["issues"]) \
                or "<li class='ok'>No issues found</li>"
            return f"""<div class="card"><div class="card-head"><h3>{esc(title)}</h3>
              <span class="badge" style="background:{gcolor(data['grade'])}">
              {data['score']}/100 &middot; {data['grade']}</span></div>
              <ul>{issues}</ul></div>"""

        rec_rows = "".join(
            f"<tr><td><span class='sev' style='background:{_SEV_COLOR.get(r['severity'],'#6b7280')}'>"
            f"{esc(r['severity'].upper())}</span></td><td>{esc(r['area'])}</td>"
            f"<td>{esc(r['issue'])}</td><td>{esc(r['recommendation'])}</td></tr>"
            for r in result.get("recommendations", [])
        ) or "<tr><td colspan='4'>No recommendations - everything looks good.</td></tr>"

        tech_rows = "".join(
            f"<tr><td>{esc(t.name)}</td><td>{esc(t.category)}</td>"
            f"<td>{esc(t.version or '-')}</td><td>{t.confidence:.2f}</td></tr>"
            for t in tech.technologies
        ) or "<tr><td colspan='4'>None detected</td></tr>"

        # Phase 2E: only rendered when --auth-detection was enabled, so a
        # default-path report's HTML is byte-for-byte unchanged.
        auth_section = ""
        if tech.authentication is not None:
            auth = tech.authentication
            auth_body_rows = "".join(
                f"<tr><td>{esc(_AUTH_CATEGORY_LABELS.get(cat, cat))}</td>"
                f"<td>{esc(f['name'])}</td><td>{f['confidence']:.2f}</td>"
                f"<td>{esc('+'.join(f['sources']))}</td></tr>"
                for cat in _AUTH_CATEGORY_ORDER
                for f in (auth.get(cat) or [])
            ) or "<tr><td colspan='4'>No authentication architecture detected</td></tr>"
            auth_section = f"""
  <div class="section-title">Authentication ({auth['total_findings']})</div>
  <table><thead><tr><th>Category</th><th>Name</th><th>Confidence</th><th>Source</th></tr></thead>
  <tbody>{auth_body_rows}</tbody></table>"""

        p = overall["parts"]
        return f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>WebIntelPro Report - {esc(result['url'])}</title>
<style>
  :root {{ font-family:-apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif; }}
  body {{ margin:0; background:#0f172a; color:#e2e8f0; }}
  .wrap {{ max-width:1000px; margin:0 auto; padding:32px 20px 64px; }}
  header {{ border-bottom:1px solid #1e293b; padding-bottom:16px; margin-bottom:24px; }}
  h1 {{ font-size:20px; margin:0 0 4px; }}
  .url {{ color:#94a3b8; font-size:13px; word-break:break-all; }}
  .hero {{ display:flex; align-items:center; gap:24px; margin:20px 0 8px; }}
  .score {{ font-size:52px; font-weight:700; line-height:1; color:{gcolor(overall['grade'])}; }}
  .parts {{ display:flex; gap:18px; flex-wrap:wrap; font-size:13px; color:#cbd5e1; }}
  .parts b {{ color:#f1f5f9; }}
  .grid {{ display:grid; grid-template-columns:1fr 1fr; gap:16px; }}
  .card {{ background:#1e293b; border:1px solid #334155; border-radius:12px; padding:16px; }}
  .card-head {{ display:flex; justify-content:space-between; align-items:center; }}
  .card h3 {{ margin:0; font-size:15px; }}
  .badge {{ color:#fff; font-size:12px; font-weight:600; padding:3px 10px; border-radius:999px; }}
  ul {{ margin:12px 0 0; padding-left:18px; font-size:13px; color:#cbd5e1; }}
  li {{ margin:4px 0; }} li.ok {{ color:#4ade80; list-style:none; margin-left:-18px; }}
  table {{ width:100%; border-collapse:collapse; margin-top:10px; font-size:13px; }}
  th,td {{ text-align:left; padding:7px 10px; border-bottom:1px solid #334155; vertical-align:top; }}
  th {{ color:#94a3b8; font-weight:600; }}
  .sev {{ color:#fff; font-size:11px; font-weight:700; padding:2px 8px; border-radius:6px; }}
  .section-title {{ margin:28px 0 6px; font-size:15px; color:#f1f5f9; }}
  footer {{ margin-top:32px; color:#64748b; font-size:12px; }}
</style></head><body><div class="wrap">
  <header><h1>WebIntelPro Enterprise X</h1><div class="url">{esc(result['url'])}</div></header>
  <div class="hero"><div class="score">{overall['score']}</div><div>
    <div style="font-size:14px;color:#94a3b8;">Overall score &middot; grade {overall['grade']}</div>
    <div class="parts" style="margin-top:8px;">
      <span>SEO <b>{p['seo']}</b></span><span>Security <b>{p['security']}</b></span>
      <span>Performance <b>{p['performance']}</b></span><span>Accessibility <b>{p['accessibility']}</b></span>
    </div></div></div>
  <div class="grid">
    {card("Technical SEO", result['seo'])}{card("Security", result['security'])}
    {card("Performance", result['performance'])}{card("Accessibility", result['accessibility'])}
  </div>
  <div class="section-title">Prioritized recommendations</div>
  <table><thead><tr><th>Severity</th><th>Area</th><th>Issue</th><th>Recommendation</th></tr></thead>
  <tbody>{rec_rows}</tbody></table>
  <div class="section-title">Detected technologies ({tech.total_detected})</div>
  <table><thead><tr><th>Technology</th><th>Category</th><th>Version</th><th>Confidence</th></tr></thead>
  <tbody>{tech_rows}</tbody></table>{auth_section}
  <footer>Generated by WebIntelPro Enterprise X on {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}</footer>
</div></body></html>"""

    def save_html(self, result: dict, path: str) -> None:
        with open(path, "w", encoding="utf-8") as f:
            f.write(self.to_html(result))

    # ------------------------------------------------------------------ excel
    def save_excel(self, result: dict, path: str) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        head = Font(bold=True, color="FFFFFF")
        fill = PatternFill("solid", fgColor="1E293B")

        def header_row(ws, cols):
            ws.append(cols)
            for cell in ws[1]:
                cell.font = head
                cell.fill = fill

        # Summary
        ws = wb.active
        ws.title = "Summary"
        header_row(ws, ["Metric", "Value"])
        o = result["overall"]
        rows = [("URL", result["url"]), ("Overall score", o["score"]),
                ("Grade", o["grade"]), ("SEO", o["parts"]["seo"]),
                ("Security", o["parts"]["security"]),
                ("Performance", o["parts"]["performance"]),
                ("Accessibility", o["parts"]["accessibility"]),
                ("Technologies detected", result["technology"].total_detected)]
        for r in rows:
            ws.append(list(r))

        # Technologies
        ws = wb.create_sheet("Technologies")
        header_row(ws, ["Name", "Category", "Version", "Confidence"])
        for t in result["technology"].technologies:
            ws.append([t.name, t.category, t.version or "-", t.confidence])

        # Recommendations
        ws = wb.create_sheet("Recommendations")
        header_row(ws, ["Severity", "Area", "Issue", "Recommendation"])
        for r in result.get("recommendations", []):
            ws.append([r["severity"], r["area"], r["issue"], r["recommendation"]])

        # Authentication (Phase 2E) -- only added when --auth-detection was
        # enabled, so a default-path workbook's sheet set is unchanged.
        auth = result["technology"].authentication
        if auth is not None:
            ws = wb.create_sheet("Authentication")
            header_row(ws, ["Category", "Name", "Confidence", "Source", "Evidence"])
            for cat in _AUTH_CATEGORY_ORDER:
                for f in auth.get(cat) or []:
                    ws.append([_AUTH_CATEGORY_LABELS.get(cat, cat), f["name"],
                              f["confidence"], "+".join(f["sources"]),
                              ", ".join(f.get("evidence", []))])

        for sheet in wb.worksheets:
            for col in sheet.columns:
                width = max((len(str(c.value)) for c in col if c.value), default=10)
                sheet.column_dimensions[col[0].column_letter].width = min(width + 2, 70)

        wb.save(path)

    # -------------------------------------------------------------------- pdf
    def save_pdf(self, result: dict, path: str) -> None:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                        Table, TableStyle)

        styles = getSampleStyleSheet()
        small = ParagraphStyle("small", parent=styles["Normal"], fontSize=8, leading=10)
        doc = SimpleDocTemplate(path, pagesize=A4, title="WebIntelPro Report",
                                leftMargin=18*mm, rightMargin=18*mm,
                                topMargin=16*mm, bottomMargin=16*mm)
        o = result["overall"]
        el = []
        el.append(Paragraph("WebIntelPro Enterprise X", styles["Title"]))
        el.append(Paragraph(result["url"], styles["Normal"]))
        el.append(Spacer(1, 8))
        el.append(Paragraph(
            f"<b>Overall score: {o['score']}/100 (grade {o['grade']})</b> &nbsp; "
            f"SEO {o['parts']['seo']} | Security {o['parts']['security']} | "
            f"Performance {o['parts']['performance']} | Accessibility {o['parts']['accessibility']}",
            styles["Normal"]))
        el.append(Spacer(1, 12))

        # Recommendations table
        el.append(Paragraph("Prioritized recommendations", styles["Heading2"]))
        data = [["Severity", "Area", "Issue", "Recommendation"]]
        for r in result.get("recommendations", []):
            data.append([r["severity"].upper(), r["area"],
                         Paragraph(r["issue"], small),
                         Paragraph(r["recommendation"], small)])
        if len(data) == 1:
            data.append(["-", "-", "No recommendations", "Everything looks good."])
        t = Table(data, colWidths=[22*mm, 24*mm, 55*mm, 65*mm], repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        el.append(t)
        el.append(Spacer(1, 14))

        # Technologies table
        el.append(Paragraph(
            f"Detected technologies ({result['technology'].total_detected})",
            styles["Heading2"]))
        tdata = [["Name", "Category", "Version", "Conf."]]
        for tech in result["technology"].technologies:
            tdata.append([tech.name, tech.category, tech.version or "-",
                          f"{tech.confidence:.2f}"])
        tt = Table(tdata, colWidths=[55*mm, 45*mm, 40*mm, 20*mm], repeatRows=1)
        tt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
        ]))
        el.append(tt)
        el.append(Spacer(1, 12))

        # Authentication (Phase 2E) -- only rendered when --auth-detection
        # was enabled, so a default-path PDF is unchanged.
        auth = result["technology"].authentication
        if auth is not None:
            el.append(Paragraph(
                f"Authentication ({auth['total_findings']})", styles["Heading2"]))
            adata = [["Category", "Name", "Conf.", "Source"]]
            for cat in _AUTH_CATEGORY_ORDER:
                for f in auth.get(cat) or []:
                    adata.append([_AUTH_CATEGORY_LABELS.get(cat, cat), f["name"],
                                  f"{f['confidence']:.2f}", "+".join(f["sources"])])
            if len(adata) == 1:
                adata.append(["-", "-", "-", "No authentication architecture detected"])
            at = Table(adata, colWidths=[35*mm, 45*mm, 15*mm, 65*mm], repeatRows=1)
            at.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
            ]))
            el.append(at)
            el.append(Spacer(1, 12))

        el.append(Paragraph(
            "Generated by WebIntelPro Enterprise X on "
            + datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"), small))
        doc.build(el)

    # ============================================================ comparison
    def comparison_console_str(self, cmp: dict) -> str:
        L = []
        bar = "=" * 72
        L.append(bar)
        L.append("  WebIntelPro Enterprise X  -  Competitor Comparison")
        L.append(bar)
        L.append(f"Primary: {cmp['primary']}")
        L.append("")
        header = f"{'SITE':32} {'OVR':>4} {'SEO':>4} {'SEC':>4} {'PERF':>4} {'A11Y':>4} {'TECH':>5}"
        L.append(header)
        L.append("-" * 72)
        for s in cmp["sites"]:
            tag = "*" if s.get("primary") else " "
            if s["status"] != "ok":
                L.append(f"{tag}{s['url'][:31]:31} {s['status']}")
                continue
            L.append(f"{tag}{s['url'][:31]:31} {s['overall']:>4} {s['seo']:>4} "
                     f"{s['security']:>4} {s['performance']:>4} "
                     f"{s['accessibility']:>4} {s['tech_count']:>5}")
        L.append("-" * 72)
        win = cmp.get("winner")
        L.append(f"Overall winner: {win}")
        L.append("")
        L.append("Per-dimension winners:")
        for dim, data in cmp["dimensions"].items():
            L.append(f"  {dim:14}: {data['winner']}")

        gap = cmp["tech_gap"]
        L.append("")
        L.append("Technology gap (primary vs competitors):")
        L.append(f"  Missing (competitors have, you don't): "
                 f"{', '.join(gap['missing']) or 'none'}")
        L.append(f"  Unique to you: {', '.join(gap['unique']) or 'none'}")
        L.append(bar)
        return "\n".join(L)

    def comparison_console(self, cmp: dict) -> None:
        print(self.comparison_console_str(cmp))

    def comparison_json(self, cmp: dict) -> str:
        return json.dumps(cmp, indent=2)

    def save_comparison_json(self, cmp: dict, path: str) -> None:
        with open(path, "w", encoding="utf-8") as f:
            f.write(self.comparison_json(cmp))

    def save_comparison_html(self, cmp: dict, path: str) -> None:
        esc = lambda x: _html.escape(str(x))
        ok = [s for s in cmp["sites"] if s["status"] == "ok"]

        def cell(site, dim):
            val = site.get(dim, 0)
            best = cmp["dimensions"].get(dim, {}).get("winner") == site["url"]
            style = "font-weight:700;color:#4ade80;" if best else ""
            return f"<td style='{style}'>{esc(val)}</td>"

        rows = ""
        for s in cmp["sites"]:
            name = esc(s["url"]) + (" <span class='pri'>primary</span>" if s.get("primary") else "")
            if s["status"] != "ok":
                rows += f"<tr><td>{name}</td><td colspan='6'>{esc(s['status'])}</td></tr>"
                continue
            rows += ("<tr><td>" + name + "</td>"
                     + cell(s, "overall") + cell(s, "seo") + cell(s, "security")
                     + cell(s, "performance") + cell(s, "accessibility")
                     + f"<td>{esc(s['tech_count'])}</td></tr>")

        gap = cmp["tech_gap"]
        missing = "".join(f"<li>{esc(t)}</li>" for t in gap["missing"]) or "<li class='ok'>none</li>"
        unique = "".join(f"<li>{esc(t)}</li>" for t in gap["unique"]) or "<li class='ok'>none</li>"

        doc = f"""<!doctype html><html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>WebIntelPro Comparison - {esc(cmp['primary'])}</title>
<style>
  :root {{ font-family:-apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif; }}
  body {{ margin:0; background:#0f172a; color:#e2e8f0; }}
  .wrap {{ max-width:1000px; margin:0 auto; padding:32px 20px 64px; }}
  h1 {{ font-size:20px; }} h2 {{ font-size:15px; margin-top:28px; }}
  .win {{ color:#4ade80; font-weight:700; }}
  table {{ width:100%; border-collapse:collapse; margin-top:12px; font-size:13px; }}
  th,td {{ text-align:left; padding:8px 10px; border-bottom:1px solid #334155; }}
  th {{ color:#94a3b8; }}
  .pri {{ background:#1d4ed8; color:#fff; font-size:10px; padding:1px 6px; border-radius:6px; }}
  .cols {{ display:grid; grid-template-columns:1fr 1fr; gap:16px; }}
  .card {{ background:#1e293b; border:1px solid #334155; border-radius:12px; padding:16px; }}
  ul {{ margin:8px 0 0; padding-left:18px; font-size:13px; }} li.ok {{ color:#4ade80; list-style:none; margin-left:-18px; }}
  footer {{ margin-top:28px; color:#64748b; font-size:12px; }}
</style></head><body><div class="wrap">
  <h1>Competitor Comparison</h1>
  <div style="color:#94a3b8;font-size:13px;">Primary: {esc(cmp['primary'])} &nbsp;&middot;&nbsp;
     Overall winner: <span class="win">{esc(cmp.get('winner'))}</span></div>
  <table><thead><tr><th>Site</th><th>Overall</th><th>SEO</th><th>Security</th>
    <th>Performance</th><th>Accessibility</th><th>Tech</th></tr></thead>
    <tbody>{rows}</tbody></table>
  <h2>Technology gap</h2>
  <div class="cols">
    <div class="card"><b>Competitors have, you don't</b><ul>{missing}</ul></div>
    <div class="card"><b>Unique to you</b><ul>{unique}</ul></div>
  </div>
  <footer>Generated by WebIntelPro Enterprise X on {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}</footer>
</div></body></html>"""
        with open(path, "w", encoding="utf-8") as f:
            f.write(doc)

    # ============================================================ site crawl
    def site_console_str(self, site: dict) -> str:
        L = []
        bar = "=" * 72
        L.append(bar)
        L.append("  WebIntelPro Enterprise X  -  Site-wide Report")
        L.append(bar)
        L.append(f"Start URL   : {site['start_url']}")
        L.append(f"Pages       : {site['pages_ok']} analyzed / {site['pages_crawled']} crawled")
        a = site["averages"]
        L.append("")
        L.append(f"SITE AVERAGES  Overall {a['overall']}  |  SEO {a['seo']}  |  "
                 f"Security {a['security']}  |  Perf {a['performance']}  |  A11y {a['accessibility']}")
        L.append("")
        L.append(f"{'PAGE':44} {'OVR':>4} {'SEO':>4} {'SEC':>4} {'PERF':>4} {'A11Y':>4}")
        L.append("-" * 72)
        for p in site["pages"]:
            if p.get("status") != "ok":
                L.append(f"{p['url'][:44]:44} {p.get('status')}")
                continue
            L.append(f"{p['url'][:44]:44} {p['overall']:>4} {p['seo']:>4} "
                     f"{p['security']:>4} {p['performance']:>4} {p['accessibility']:>4}")
        L.append("-" * 72)
        w = site["worst"].get("overall")
        if w:
            L.append(f"Weakest page: {w['url']} ({w['score']}/100)")
        L.append(f"Technologies across site ({len(site['technologies'])}): "
                 f"{', '.join(site['technologies'][:25])}"
                 + (" ..." if len(site['technologies']) > 25 else ""))
        L.append("")
        L.append("TOP RECURRING ISSUES")
        for r in site["recommendations"][:10]:
            L.append(f"  [{r['severity'].upper():8}] {r['area']}: {r['issue']}"
                     f"  (on {r['pages_affected']} page(s))")
        L.append(bar)
        return "\n".join(L)

    def site_console(self, site: dict) -> None:
        print(self.site_console_str(site))

    def save_site_json(self, site: dict, path: str) -> None:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(site, f, indent=2)

    def save_site_html(self, site: dict, path: str) -> None:
        esc = lambda x: _html.escape(str(x))
        a = site["averages"]
        rows = ""
        for p in site["pages"]:
            if p.get("status") != "ok":
                rows += f"<tr><td>{esc(p['url'])}</td><td colspan='5'>{esc(p.get('status'))}</td></tr>"
                continue
            rows += (f"<tr><td>{esc(p['url'])}</td><td>{p['overall']}</td><td>{p['seo']}</td>"
                     f"<td>{p['security']}</td><td>{p['performance']}</td><td>{p['accessibility']}</td></tr>")
        recs = "".join(
            f"<tr><td>{esc(r['severity'].upper())}</td><td>{esc(r['area'])}</td>"
            f"<td>{esc(r['issue'])}</td><td>{r['pages_affected']}</td></tr>"
            for r in site["recommendations"][:25]) or "<tr><td colspan='4'>None</td></tr>"
        doc = f"""<!doctype html><html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>WebIntelPro Site Report - {esc(site['start_url'])}</title>
<style>
 body{{margin:0;background:#0f172a;color:#e2e8f0;font-family:-apple-system,Segoe UI,Roboto,Arial,sans-serif;}}
 .wrap{{max-width:1000px;margin:0 auto;padding:32px 20px 64px;}} h1{{font-size:20px;}}
 h2{{font-size:15px;margin-top:26px;}} table{{width:100%;border-collapse:collapse;margin-top:10px;font-size:13px;}}
 th,td{{text-align:left;padding:7px 10px;border-bottom:1px solid #334155;}} th{{color:#94a3b8;}}
 .avg{{font-size:14px;color:#cbd5e1;margin-top:8px;}} .avg b{{color:#f1f5f9;}}
 footer{{margin-top:28px;color:#64748b;font-size:12px;}}
</style></head><body><div class="wrap">
 <h1>Site-wide Report</h1>
 <div style="color:#94a3b8;font-size:13px;">{esc(site['start_url'])} &middot; {site['pages_ok']} pages analyzed</div>
 <div class="avg">Averages &middot; Overall <b>{a['overall']}</b> | SEO <b>{a['seo']}</b> |
   Security <b>{a['security']}</b> | Performance <b>{a['performance']}</b> | Accessibility <b>{a['accessibility']}</b></div>
 <h2>Pages</h2>
 <table><thead><tr><th>Page</th><th>Overall</th><th>SEO</th><th>Security</th><th>Perf</th><th>A11y</th></tr></thead>
 <tbody>{rows}</tbody></table>
 <h2>Recurring issues</h2>
 <table><thead><tr><th>Severity</th><th>Area</th><th>Issue</th><th>Pages</th></tr></thead>
 <tbody>{recs}</tbody></table>
 <footer>Generated by WebIntelPro Enterprise X on {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}</footer>
</div></body></html>"""
        with open(path, "w", encoding="utf-8") as f:
            f.write(doc)
